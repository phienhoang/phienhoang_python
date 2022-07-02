import openpyxl


def get_value_excel(filename, cellname):
    wb = openpyxl.load_workbook(filename)
    Sheet1 = wb['Sheet1']
    wb.close()
    return Sheet1[cellname].value

def get_value_excel_2(filename, row, column):
    wb = openpyxl.load_workbook(filename)
    Sheet1 = wb['Sheet1']
    wb.close()
    return Sheet1.cell(row=row, column=column).value


def update_value_excel(filename, cellname, value):
    wb = openpyxl.load_workbook(filename)
    Sheet1 = wb['Sheet1']
    Sheet1[cellname].value = value
    wb.close()
    wb.save(filename)

def update_value_excel_2(filename, row, column, value):
    wb = openpyxl.load_workbook(filename)
    Sheet1 = wb['Sheet1']
    Sheet1.cell(row=row, column=column).value = value
    wb.close()
    wb.save(filename)

filename = 'data/fileexcel.xlsx'
cellname = 'G6'
new_value = '2k'
update_value_excel(filename,cellname,new_value)
bien_x = get_value_excel(filename, cellname)
print('bien_x',bien_x)
#
# col_name_acc = 'F'
# col_name_pass = 'G'
# filename = 'file.xlsx'
# for i_row in range(7,18):
#     cell_name_acc = "%s%s"%(col_name_acc,i_row)
#     cell_name_pass= "%s%s"%(col_name_pass,i_row)
#     account = get_value_excel(filename,cell_name_acc)
#     password = get_value_excel(filename,cell_name_pass)
#     print('current account',account)
#     print('current password',password)