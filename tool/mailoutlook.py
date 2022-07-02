import openpyxl
import os
import imaplib
import email
from email.header import decode_header
t1 = 0.7
t2 = 1
t3 = 0.1
t4 = 0.2

def get_value_excel_2(filename, row, column):
    wb = openpyxl.load_workbook(filename)
    Sheet1 = wb['Sheet1']
    wb.close()
    return Sheet1.cell(row=row, column=column).value


def update_value_excel_2(filename, row, column, value):
    wb = openpyxl.load_workbook(filename)
    Sheet1 = wb['Sheet1']
    Sheet1.cell(row=row, column=column).value = value
    wb.close()
    wb.save(filename)


TEN = ["James", "Robert", "John", "Michael", "William", "David", "Richard", "Joseph", "Thomas", "Charles", "Christopher", "Daniel", "Matthew", "Anthony", "Mark", "Donald", "Steven", "Paul", "Andrew", "Joshua", "Kenneth", "Kevin", "Brian", "George", "Edward", "Ronald", "Timothy", "Jason", "Jeffrey", "Ryan", "Jacob", "Gary", "Nicholas", "Eric", "Jonathan", "Stephen", "Larry", "Justin", "Scott", "Brandon", "Benjamin", "Samuel", "Gregory", "Frank", "Alexander", "Raymond", "Patrick", "Jack", "Dennis", "Jerry", "Tyler", "Aaron", "Jose", "Adam", "Henry", "Nathan", "Douglas", "Zachary", "Peter", "Kyle", "Walter", "Ethan", "Jeremy", "Harold", "Keith", "Christian", "Roger", "Noah", "Gerald", "Carl", "Terry", "Sean", "Austin", "Arthur", "Lawrence", "Jesse", "Dylan", "Bryan", "Joe", "Jordan", "Billy", "Bruce", "Albert", "Willie", "Gabriel", "Logan", "Alan", "Juan", "Wayne", "Roy", "Ralph", "Randy", "Eugene", "Vincent", "Russell", "Elijah", "Louis", "Bobby", "Philip", "Johnny"]
HO = ["Johnson" ,"Williams" ,"Brown" ,"Jones" ,"Garcia" ,"Miller" ,"Davis" ,"Rodriguez" ,"Martinez" ,"Hernandez" ,"Lopez" ,"Gonzalez" ,"Wilson" ,"Anderson" ,"Thomas" ,"Taylor" ,"Moore" ,"Jackson" ,"Martin" ,"Lee" ,"Perez" ,"Thompson" ,"White" ,"Harris" ,"Sanchez" ,"Clark" ,"Ramirez" ,"Lewis" ,"Robinson" ,"Walker" ,"Young" ,"Allen" ,"King" ,"Wright" ,"Scott" ,"Torres" ,"Nguyen" ,"Hill" ,"Flores" ,"Green" ,"Adams" ,"Nelson" ,"Baker" ,"Hall" ,"Rivera" ,"Campbell" ,"Mitchell" ,"Carter" ,"Roberts"]
filename = 'data/accmail.xlsx'

# tim dong cuoi cung
ws = openpyxl.load_workbook(filename)
Sheet = ws['Sheet1']
column = Sheet['A']
size = len(column)+1

for lanchay in range(size,size+1):
    os.system('windscribe connect us')
    from driver_chrome import *
    driver.delete_all_cookies()
    ten = random.choice(TEN)
    ho = random.choice(HO)
    acc = ten + ho + random.choice('abcdefghijklmnopqrstusvwyz') +random.choice('abcdefghijklmnopqrstusvwyz')+random.choice('abcdefghijklmnopqrstusvwyz')+random.choice('abcdefghijklmnopqrstusvwyz')+ random.choice('123456789')
    mk = ho+ ten + random.choice('abcdefghijklmnopqrstusvwyz') + random.choice('123456789') 
    driver.get('https://outlook.live.com')
    sleep(random.uniform(t1,t2))
    driver.find_element_by_xpath('/html/body/header/div/div/section/div/div[2]/a').click()
    mail = driver.find_element_by_id('MemberName')
    actions.move_to_element(mail).click(mail).perform()
    for el in acc:
        mail.send_keys(el)
        sleep(random.uniform(t3,t4))
    driver.find_element_by_id('iSignupAction').click()
    driver.implicitly_wait(7)
    while True:
        try:
            if driver.find_element_by_id('MemberNameError').text == 'Someone already has this email address. Try another name or ':
                mail = driver.find_element_by_id('MemberName')
                for el in acc:
                    mail.send_keys(el)
                    sleep(random.uniform(t3,t4))
                driver.find_element_by_id('iSignupAction').click()
            else:
                break
        except:
            break
    sleep(random.uniform(t1,t2))
    password1 = driver.find_element_by_id('PasswordInput')
    actions = webdriver.ActionChains(driver)
    actions.move_to_element(password1).click(password1).perform()
    for el in mk:
        password1.send_keys(el)
        sleep(random.uniform(t3,t4))
    driver.find_element_by_id('iOptinEmail').click()
    driver.find_element_by_id('iSignupAction').click()
    sleep(random.uniform(t1,t2))
    Ten1 = driver.find_element_by_id('FirstName')
    actions = webdriver.ActionChains(driver)
    actions.move_to_element(Ten1).click(Ten1).perform()
    for el in ten:
        Ten1.send_keys(el)
        sleep(random.uniform(t3,t4))
    Ho1 = driver.find_element_by_id('LastName')
    actions = webdriver.ActionChains(driver)
    actions.move_to_element(Ho1).click(Ho1).perform()
    for el in ho:
        Ho1.send_keys(el)
        sleep(random.uniform(t3,t4))
    driver.find_element_by_id('iSignupAction').click()
    sleep(random.uniform(t1,t2))

    qgia = Select(driver.find_element_by_id('Country'))
    qgia.select_by_value('US')
    sleep(random.uniform(t1,t2))

    month1 = Select(driver.find_element_by_id('BirthMonth'))
    month1.select_by_value(str(random.randint(1,12)))
    sleep(random.uniform(t1,t2))
    day1 = Select(driver.find_element_by_id('BirthDay'))
    day1.select_by_value(str(random.randint(1,28)))
    sleep(random.uniform(t1,t2))

    year1 = driver.find_element_by_id('BirthYear')
    nam = str(random.randint(1980,2001))
    for el in nam:
    	year1.send_keys(el)
    	sleep(random.uniform(t3,t4))
    sleep(random.uniform(t1,t2))

    driver.find_element_by_id('iSignupAction').click()
    driver.implicitly_wait(600)
    element = driver.find_element_by_xpath('//div[@title="Inbox"]')
    print('Xong')
    

    update_value_excel_2(filename,lanchay,1,acc+'@outlook.com')
    update_value_excel_2(filename,lanchay,2,mk)

    # driver.get('https://facebook.com')
    # driver.implicitly_wait(30)
    # sleep(random.uniform(t1,t2))
    # driver.find_element_by_xpath('/html/body/div[1]/div[2]/div[1]/div/div/div/div[2]/div/div[1]/form/div[5]/a').click()
    # sleep(random.uniform(t1,t2))
    # Ten2 = driver.find_element_by_name("firstname")
    # actions = webdriver.ActionChains(driver)
    # actions.move_to_element(Ten2)
    # actions.click(Ten2)
    # actions.perform()
    # for el in ten:
    #     Ten2.send_keys(el)
    #     sleep(random.uniform(t3,t4))
    # Ho2 = driver.find_element_by_name('lastname')
    # actions = webdriver.ActionChains(driver)
    # actions.move_to_element(Ho2)
    # actions.click(Ho2)
    # actions.perform()
    # for el in ho:
    #     Ho2.send_keys(el)
    #     sleep(random.uniform(t3,t4))
    # mail1 = driver.find_element_by_name('reg_email__')
    # actions = webdriver.ActionChains(driver)
    # actions.move_to_element(mail1)
    # actions.click(mail1)
    # actions.perform()
    # acc = acc + '@outlook.com'
    # for el in acc:
    #     mail1.send_keys(el)
    #     sleep(random.uniform(t3,t4))
    # mail2 = driver.find_element_by_name('reg_email_confirmation__')
    # actions = webdriver.ActionChains(driver)
    # actions.move_to_element(mail2).click(mail2).perform()
    # for el in acc:
    #     mail2.send_keys(el)
    #     sleep(random.uniform(t3,t4))
    # password2 = driver.find_element_by_name('reg_passwd__')
    # actions = webdriver.ActionChains(driver)
    # actions.move_to_element(password2).click(password2).perform()
    # for el in mk:
    #     password2.send_keys(el)
    #     sleep(random.uniform(t3,t4))
    # ngay = driver.find_element_by_id('day')
    # actions = webdriver.ActionChains(driver)
    # actions.move_to_element(ngay)
    # actions.click(ngay)
    # actions.perform()
    # sleep(random.uniform(t1,t2))
    # select1 = Select(driver.find_element_by_id('day'))
    # day = select1.select_by_value(str(random.randint(1,28)))
    # actions = webdriver.ActionChains(driver)
    # actions.move_to_element(day)
    # actions.click(day)
    # actions.perform()
    # sleep(random.uniform(t1,t2))
    # thang = driver.find_element_by_id('month')
    # actions = webdriver.ActionChains(driver)
    # actions.move_to_element(thang).click(thang).perform()
    # select2 = Select(driver.find_element_by_id('month'))
    # month = select2.select_by_value(str(random.randint(1,12)))
    # actions = webdriver.ActionChains(driver)
    # actions.move_to_element(month).click(month).perform()
    # sleep(random.uniform(t1,t2))
    # nam = driver.find_element_by_id('year')
    # actions = webdriver.ActionChains(driver)
    # actions.move_to_element(nam).click(nam).perform()
    # sleep(random.uniform(t1,t2))
    # select3 = Select(driver.find_element_by_id('year'))
    # year = select3.select_by_value(str(random.randint(1980,2000)))
    # actions = webdriver.ActionChains(driver)
    # actions.move_to_element(year).click(year).perform()
    # sleep(random.uniform(t1,t2))
    # sex = driver.find_elements_by_name('sex')
    # sex_ = random.choice(sex[:2])
    # actions = webdriver.ActionChains(driver)
    # actions.move_to_element(sex_).click(sex_).perform()
    # # nu = driver.find_element_by_xpath('''//label[contains(text(), "Женщина")]''')
    # # nam = driver.find_element_by_xpath('/html/body/div[3]/div[2]/div/div/div[2]/div/div/div[1]/form/div[1]/div[7]/span/span[2]/input')
    # # random.choice([nam,nu]).click()
    # sleep(random.uniform(t1,t2))
    # driver.find_element_by_name('websubmit').click()
    # sleep(random.uniform(t1,t2))
    # # driver.find_element_by_xpath('/html/body/div[1]/div[3]/div[1]/div/div[1]/div/div[1]/div/div/div/div/div/div/div/div/div/div[2]/div/div/div[2]/div/div/div/div/div[1]/div/span/span').click()
    # driver.implicitly_wait(600)
    # # driver.find_element_by_xpath('/html/body/div[1]/div[3]/div[1]/div/div[1]/div/div[1]/div/div/div/div/div/div/div/div/div/div[3]/div/div/div/div/div[1]/div/span').click()
    # # driver.implicitly_wait(30)

    # username = acc
    # passw = mk
    # '''def clean(text):
    #     # clean text for creating a folder
    #     return "".join(c if c.isalnum() else "_" for c in text)'''
    # # create an IMAP4 class with SSL 
    # imap = imaplib.IMAP4_SSL("imap.outlook.com")
    # # authenticate
    # imap.login(username, passw)
    # status, messages = imap.select("INBOX")
    # # number of top emails to fetch
    # N = 10
    # nguoigui = '"Facebook" <registration@facebookmail.com>'
    # tieude = 'ваш код подтверждения для Facebook'
    # # total number of emails
    # messages = int(messages[0])
    # for i in range(messages, messages-N, -1):
    #     # fetch the email message by ID
    #     res, msg = imap.fetch(str(i), "(RFC822)")
    #     body = ""
    #     for response in msg:
    #         if isinstance(response, tuple):
    #             # parse a bytes email into a message object
    #             msg = email.message_from_bytes(response[1])
    #             # decode the email subject
    #             subject, encoding = decode_header(msg["Subject"])[0]
    #             if isinstance(subject, bytes):
    #                 # if it's a bytes, decode to str
    #                 subject = subject.decode(encoding)
    #             # decode email sender
    #             From, encoding = decode_header(msg.get("From"))[0]
    #             if isinstance(From, bytes):
    #                 From = From.decode(encoding)
    #             if (From == nguoigui) and (tieude in subject):
    #                 # print("Subject:", subject)
    #                 # print("From:", From)
    #                 # if the email message is multipart
    #                 if msg.is_multipart():
    #                     # iterate over email parts
    #                     for part in msg.walk():
    #                         # extract content type of email
    #                         content_type = part.get_content_type()
    #                         content_disposition = str(part.get("Content-Disposition"))
    #                         try:
    #                             # get the email body
    #                             body = part.get_payload(decode=True).decode()
    #                             # print(body)
    #                             break
    #                         except:
    #                             pass
    #                         '''if content_type == "text/plain" and "attachment" not in content_disposition:
    #                             # print text/plain emails and skip attachments
    #                             print(body)
    #                         elif "attachment" in content_disposition:
    #                             # download attachment
    #                             filename = part.get_filename()
    #                             if filename:
    #                                 folder_name = clean(subject)
    #                                 if not os.path.isdir(folder_name):
    #                                     # make a folder for this email (named after the subject)
    #                                     os.mkdir(folder_name)
    #                                 filepath = os.path.join(folder_name, filename)
    #                                 # download attachment and save it
    #                                 open(filepath, "wb").write(part.get_payload(decode=True))'''
    #                 else:
    #                     # extract content type of email
    #                     content_type = msg.get_content_type()
    #                     # get the email body
    #                     body = msg.get_payload(decode=True).decode()
    #                     if content_type == "text/plain":
    #                         # print only text email parts
    #                         # print(body)
    #                         break
    #                 '''if content_type == "text/html":
    #                     # if it's HTML, create a new HTML file and open it in browser
    #                     folder_name = clean(subject)
    #                     if not os.path.isdir(folder_name):
    #                         # make a folder for this email (named after the subject)
    #                         os.mkdir(folder_name)
    #                     filename = "index.html"
    #                     filepath = os.path.join(folder_name, filename)
    #                     # write the file
    #                     open(filepath, "w").write(body)
    #                 print("="*100)'''
    #     if body:
    #         break
    # vitri = body.index('FB-')
    # MA = body[vitri+len('FB-'):vitri+len('FB-')+5]
    # driver.implicitly_wait(30)
    # ma = driver.find_element_by_id('code_in_cliff')

    # for el in MA:
    #     ma.send_keys(el)
    #     sleep(random.uniform(t3,t4))
    # driver.find_element_by_name('confirm').click()
    # driver.find_element_by_xpath('/html/body/div[4]/div[2]/div/div/div/div[3]/div/a').click()
    # sleep(random.uniform(t1,t2))
    # driver.get('https://www.facebook.com/phien.hoang.902')
    # sleep(random.uniform(t1,t2))
    # driver.find_element_by_xpath('/html/body/div[1]/div/div[1]/div/div[3]/div/div/div[1]/div[1]/div/div/div[3]/div/div/div/div[2]/div/div/div/div[1]/div/div/div/div[1]/div[2]/span/span').click()
    # sleep(random.uniform(t1,t2))
    # driver.find_element_by_xpath('//div[@aria-label="Like"]').click()



    # update_value_excel_2(filename,lanchay,4,acc)
    # update_value_excel_2(filename,lanchay,5,mk)